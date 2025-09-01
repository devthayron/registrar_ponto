from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required,user_passes_test
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils import timezone
from .models import RegistroPonto, Colaborador, Lider
from django.template.loader import render_to_string
from django.http import HttpResponse
from xhtml2pdf import pisa
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from datetime import timedelta,date
from django.utils.timezone import localdate
from django.http import JsonResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_http_methods
from django.core.serializers import serialize, deserialize
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.urls import reverse
import json
from .models import Colaborador, Lider, RegistroPonto
import qrcode
import io
import base64
from django.shortcuts import render
from django.http import HttpResponseBadRequest
from datetime import datetime, timedelta


# ------------------  Usuário  ------------------
User = get_user_model()

# ------------------  Gerente  ------------------
def is_gerente(user):
    return user.nivel == 'gerente'

# ------------------  Filtro em exportações  ------------------

def filtrar_registros(request):
    cpf = request.GET.get('cpf', '').strip()
    cpf_limpo = cpf.replace('.', '').replace('-', '')
    lider_id = request.GET.get('lider', '').strip()
    data_inicial = request.GET.get('data_inicial')
    data_final = request.GET.get('data_final')

    registros = RegistroPonto.objects.select_related('colaborador').all()

    if cpf_limpo and len(cpf_limpo) == 11 and cpf_limpo.isdigit():
        registros = registros.filter(colaborador__cpf=cpf_limpo)

    if lider_id:
        registros = registros.filter(colaborador__lider_id=lider_id)

    if data_inicial:
        try:
            data_ini = date.fromisoformat(data_inicial)
            registros = registros.filter(data__gte=data_ini)
        except ValueError:
            pass

    if data_final:
        try:
            data_fim = date.fromisoformat(data_final)
            registros = registros.filter(data__lte=data_fim)
        except ValueError:
            pass

    return registros.order_by('colaborador__nome', 'data')


# ------------------  Excel  ------------------
@login_required
def baixar_historico_geral_excel(request):
    registros = filtrar_registros(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico Geral"

    headers = ['CPF', 'Nome', 'Data', 'Entrada', 'Contrato',]  # 'Status'
    header_font = Font(bold=True)
    alignment = Alignment(horizontal='center')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = alignment

    for row_num, registro in enumerate(registros, start=2):
        entrada_formatada = timezone.localtime(registro.entrada).strftime('%H:%M') if registro.entrada else '---'

        ws.cell(row=row_num, column=1, value=registro.colaborador.cpf)
        
        nome = registro.colaborador.nome
        if not registro.colaborador.is_active:
            nome = f"(INATIVO) {nome}"
        ws.cell(row=row_num, column=2, value=nome)
        ws.cell(row=row_num, column=3, value=registro.data.strftime('%d/%m/%Y') if registro.data else '')
        ws.cell(row=row_num, column=4, value=entrada_formatada)
        ws.cell(row=row_num, column=5, value=getattr(registro, 'lider_nome', ''))
        # ws.cell(row=row_num, column=6, value='ATIVO' if registro.colaborador.is_active else 'INATIVO')
        
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = "historico_geral.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response

# <a href="{% url 'baixar_presenca_excel' %}?cpf={{ cpf }}&lider={{ lider }}&data_inicial={{ data_inicial }}&data_final={{ data_final }}" title="Baixar Presença Excel">
# <button type="button">📥 Baixar Presença</button>
# </a>
# baixar presenca 
# @login_required
# def baixar_presenca_excel(request):
#     registros = filtrar_registros(request)

#     data_inicial = request.GET.get('data_inicial')
#     data_final = request.GET.get('data_final')

#     if data_inicial and data_final:
#         registros = registros.filter(data__range=[data_inicial, data_final])
#     else:
#         hoje = localdate()
#         registros = registros.filter(data__month=hoje.month, data__year=hoje.year)

#     # Obter dias únicos com presença (ordenados)
#     dias_com_presenca = sorted(set(r.data.day for r in registros))

#     presencas = defaultdict(lambda: {
#         'nome': '',
#         'cpf': '',
#         'contrato': '',
#         'dias': {}
#     })

#     for r in registros:
#         dia = r.data.day
#         cpf_colaborador = r.colaborador.cpf
#         presencas[cpf_colaborador]['nome'] = r.colaborador.nome
#         presencas[cpf_colaborador]['cpf'] = cpf_colaborador
#         lider = r.colaborador.lider
#         presencas[cpf_colaborador]['contrato'] = lider.nome if lider else '—'
#         presencas[cpf_colaborador]['dias'][dia] = 'S'

#     # Criar planilha em modo write_only para performance
#     wb = Workbook(write_only=True)
#     ws = wb.create_sheet(title="Controle de Presença")

#     # Cabeçalho sem CPF
#     headers = ["Funcionário", "Contrato"] + [str(d) for d in dias_com_presenca]
#     ws.append(headers)

#     # Dados por colaborador
#     for dados in sorted(presencas.values(), key=lambda x: x['contrato']):
#         linha = [
#             dados['nome'],
#             # dados['cpf'],  # CPF desativado
#             dados['contrato'],
#         ] + [dados['dias'].get(d, '') for d in dias_com_presenca]
#         ws.append(linha)

#     # Totais por dia
#     total_por_dia = ["Total", ""]  # Coluna CPF ignorada, então só 2 fixas
#     for d in dias_com_presenca:
#         total = sum(1 for dados in presencas.values() if dados['dias'].get(d) == 'S')
#         total_por_dia.append(total)
#     ws.append(total_por_dia)

#     # Preparar resposta para download direto
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="controle_presenca.xlsx"'
#     wb.save(response)

#     return response


# ------------------  PDF  ------------------
@login_required
def baixar_historico_geral_pdf(request):
    registros = filtrar_registros(request).order_by('lider_nome', 'data', 'colaborador__nome')

    data_inicial = request.GET.get('data_inicial')
    data_final = request.GET.get('data_final')

    # Se não vieram, define como data de hoje
    hoje = localdate()
    if not data_inicial:
        data_inicial = hoje.strftime('%Y-%m-%d')
    if not data_final:
        data_final = hoje.strftime('%Y-%m-%d')

    html_string = render_to_string('ponto/pdf_pontos.html', {
        'registros': registros,
        'cpf': request.GET.get('cpf', ''),
        'lider': request.GET.get('lider', ''),
        'data_inicial': data_inicial,
        'data_final': data_final,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=historico_geral_ponto.pdf'

    pisa_status = pisa.CreatePDF(html_string, dest=response)

    if pisa_status.err:
        return HttpResponse('Erro ao gerar PDF', status=500)

    return response


# ------------------  Login  ------------------

def login_view(request):
    if request.user.is_authenticated:
        if request.user.nivel == 'gerente':
            return redirect('listar_pontos')
        else:
            return redirect('registrar_ponto')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if user.nivel == 'gerente':
                return redirect('listar_pontos')  
            else:
                return redirect('registrar_ponto')  
        else:
            messages.error(request, 'Usuário ou senha incorretos.')

    return render(request, 'ponto/login.html')

# ------------------  logout  ------------------
@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

# ------------------ Registro ------------------
@login_required
def registrar_ponto(request):
    if request.method == 'POST':
        cpf = request.POST.get('cpf', '').strip()
        cpf_limpo = cpf.replace('.', '').replace('-', '')

        if not (cpf_limpo.isdigit() and len(cpf_limpo) == 11):
            messages.error(request, 'CPF inválido. Informe exatamente 11 dígitos numéricos.')
            return redirect('registrar_ponto')

        colaborador = Colaborador.objects.filter(cpf=cpf_limpo).first()
        if not colaborador:
            messages.error(request, 'CPF não cadastrado no sistema. Contate o administrador.')
            return redirect('registrar_ponto')

        # ✅ Verifica se o colaborador está inativo
        if not colaborador.is_active:
            messages.error(request, 'Colaborador inativo, não permitido registrar.')
            return redirect('registrar_ponto')

        data_str = request.POST.get('data')
        hora_str = request.POST.get('hora')

        try:
            data = datetime.strptime(data_str, '%Y-%m-%d').date()
            hora = datetime.strptime(hora_str, '%H:%M').time()
            data_hora = datetime.combine(data, hora)
            data_hora = timezone.make_aware(data_hora)
        except (ValueError, TypeError):
            messages.error(request, 'Data ou hora inválida.')
            return redirect('registrar_ponto')

        try:
            registro, created = RegistroPonto.objects.get_or_create(
                colaborador=colaborador,
                data=data
            )

            if registro.entrada:
                messages.error(request, 'Entrada já registrada para esta data.')
                return redirect('registrar_ponto')

            intervalo = timedelta(seconds=10)
            if registro.entrada and (data_hora - registro.entrada) < intervalo:
                messages.error(request, 'Leitura ignorada: entrada já registrada recentemente.')
                return redirect('registrar_ponto')

            registro.entrada = data_hora
            registro.save()
            messages.success(request, 'Entrada registrada com sucesso!')

        except Exception as e:
            messages.error(request, f'Erro ao registrar entrada: {e}')

        return redirect('registrar_ponto')

    return render(request, 'ponto/registrar_ponto.html')

# ------------------  Listagem  ------------------
@login_required
def listar_pontos(request):
    if not is_gerente(request.user):
        messages.error(request, 'Acesso restrito. Apenas gerentes podem acessar esta página.')
        return redirect('registrar_ponto')

    cpf = request.GET.get('cpf', '').strip()
    cpf_limpo = cpf.replace('.', '').replace('-', '')
    lider_id = request.GET.get('lider', '').strip()
    data_hoje = localdate()
    data_inicial = request.GET.get('data_inicial') or data_hoje.isoformat()
    data_final = request.GET.get('data_final') or data_hoje.isoformat()


    filtros_usados = any([cpf_limpo, lider_id, data_inicial, data_final])
    registros = RegistroPonto.objects.select_related('colaborador').all()

    if cpf_limpo and len(cpf_limpo) == 11 and cpf_limpo.isdigit():
        registros = registros.filter(colaborador__cpf=cpf_limpo)

    if lider_id:
        registros = registros.filter(colaborador__lider_id=lider_id)

    if filtros_usados:
        try:
            if data_inicial:
                data_ini = date.fromisoformat(data_inicial)
                registros = registros.filter(data__gte=data_ini)
            if data_final:
                data_fim = date.fromisoformat(data_final)
                registros = registros.filter(data__lte=data_fim)
        except ValueError:
            messages.warning(request, 'Datas inválidas. Nenhum filtro de data foi aplicado.')
    else:
        data_hoje = localdate()
        registros = registros.filter(data=data_hoje)

    total_presencas = registros.count()
    
    registros = registros.order_by('-data', 'lider_nome', '-entrada')
    paginator = Paginator(registros, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    lideres = Lider.objects.all()

    return render(request, 'ponto/listar_pontos.html', {
        'registros': page_obj.object_list,
        'cpf': cpf,
        'lider': lider_id,
        'page_obj': page_obj,
        'lideres': lideres,
        'data_inicial': data_inicial,
        'data_final': data_final,
        'total_presencas': total_presencas,
        'request': request,
    })

# ------------------  Admin-exportar json  ------------------

@staff_member_required
def exportar_json_admin(request):
    dados = {
        'lideres': json.loads(serialize('json', Lider.objects.all())),
        'colaboradores': json.loads(serialize('json', Colaborador.objects.all())),
        'registros': json.loads(serialize('json', RegistroPonto.objects.all())),
    }
    response = JsonResponse(dados, safe=False)
    response['Content-Disposition'] = 'attachment; filename=backup_dados.json'
    return response


# ------------------  Admin-importar json  ------------------
@staff_member_required
@require_http_methods(["GET", "POST"])
def importar_json_admin(request):
    if request.method == "POST":
        json_file = request.FILES.get("json_file")
        if not json_file:
            messages.error(request, "Nenhum arquivo enviado.")
            return HttpResponseRedirect(reverse("admin:index"))

        try:
            dados = json.load(json_file)
            for model_key in ['lideres', 'colaboradores', 'registros']:
                for obj_data in dados.get(model_key, []):
                    for obj in deserialize("json", json.dumps([obj_data])):
                        obj.save()
            messages.success(request, "Dados importados com sucesso.")
        except Exception as e:
            messages.error(request, f"Erro ao importar: {e}")
        return HttpResponseRedirect(reverse("admin:index"))

    return render(request, "admin/importar_json.html")

@user_passes_test(is_gerente)
def formulario_view(request):
    import re

    mensagem = ''
    colaborador = None
    lideres = Lider.objects.all()

    if request.method == 'POST':
        cpf = request.POST.get('cpf')
        cpf = re.sub(r'\D', '', cpf)  # Remove pontos e traços do CPF

        nome = request.POST.get('nome')
        lider_id = request.POST.get('lider')
        is_active_raw = request.POST.get('is_active', 'ativo')

        is_active = (is_active_raw == 'ativo')
        lider = Lider.objects.filter(id=lider_id).first() if lider_id else None

        if 'buscar' in request.POST:
            colaborador = Colaborador.objects.filter(cpf=cpf).first()
            if not colaborador: 
                mensagem = 'Colaborador não encontrado.'

        elif 'cadastrar' in request.POST:
            if Colaborador.objects.filter(cpf=cpf).exists():
                mensagem = 'Este CPF já está cadastrado.'
            else:
                Colaborador.objects.create(
                    cpf=cpf,
                    nome=nome,
                    lider=lider,
                    is_active=is_active
                )
                mensagem = 'Colaborador cadastrado com sucesso!'

        elif 'editar' in request.POST:
            colaborador = Colaborador.objects.filter(cpf=cpf).first()
            if colaborador:
                colaborador.nome = nome
                colaborador.lider = lider
                colaborador.is_active = is_active
                colaborador.save()
                mensagem = 'Colaborador atualizado com sucesso!'
            else:
                mensagem = 'Colaborador não encontrado para edição.'

    return render(request, 'ponto/formulario.html', {
        'mensagem': mensagem,
        'colaborador': colaborador,
        'lideres': lideres
    })


def formulario_etiqueta(request):
    """Exibe o formulário"""
    return render(request, "ponto/formulario.html")

def gerar_etiqueta(request):
    """Gera o QR Code e exibe na página de etiqueta"""
    nome = request.GET.get('nome', '').strip()
    cpf = request.GET.get('cpf', '').strip()

    # Validação simples
    if not nome or len(cpf) != 11 or not cpf.isdigit():
        return HttpResponseBadRequest("Nome ou CPF inválido. O CPF deve ter 11 dígitos numéricos.")

    # Texto que será codificado no QR Code
    conteudo_qr = cpf

    # Criar QR Code
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(conteudo_qr)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    # Converter imagem para base64
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    # Enviar para o template
    return render(request, "ponto/etiqueta.html", {
        "nome": nome,
        "cpf": cpf,
        "qr_code": qr_base64
    })
